"""
Job fetcher: Parses Excel -> Scrapes HTML -> Gemini extraction -> Reviews -> Excel export
"""

import json
import os
import re
import requests
import pandas as pd
from bs4 import BeautifulSoup
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

load_dotenv()


class Genai:
    """Gemini API client."""

    def __init__(self, api_key=None, model=None):
        self.api_key = api_key or os.getenv('GEMINI_API_KEY')
        self.model = model or os.getenv('GEMINI_MODEL', 'gemini-2.0-flash')

        if not self.api_key:
            raise ValueError("GEMINI_API_KEY not found")

    def send_request(self, system_prompt, user_prompt, use_search=False):
        """Send request to Gemini API."""
        url = f"https://generativelanguage.googleapis.com/v1beta/models/{self.model}:generateContent"

        data = {
            "system_instruction": {"parts": [{"text": system_prompt}]},
            "contents": [{"parts": [{"text": user_prompt}]}],
        }

        if use_search:
            data["tools"] = [{"google_search": {}}]

        try:
            response = requests.post(
                url,
                headers={"x-goog-api-key": self.api_key, "Content-Type": "application/json"},
                json=data,
                timeout=60
            )
            response.raise_for_status()
            return response.json()["candidates"][0]["content"]["parts"][0]["text"]
        except Exception as e:
            return f"ERROR: {str(e)}"

    def parse_job(self, html, config):
        """Extract job data from HTML."""
        system_prompt = """You are a JSON data extractor for job postings.

Extract these fields from HTML:
- company_name, role_name, experience_required, experience_type, location (with exact and city), remote, hybrid_or_flexible, match_score (1-10 based on config)

CRITICAL: Return ONLY valid JSON. No markdown, no code blocks, no extra text.

Format:
{"company_name":"","role_name":"","experience_required":"","experience_type":"","location":{"exact":"","city":""},"remote":"","hybrid_or_flexible":"","match_score":0}"""

        user_prompt = f"HTML:\n{html}\n\nConfig:\n{json.dumps(config)}\n\nReturn JSON only:"

        return self.send_request(system_prompt, user_prompt)

    def get_reviews(self, company_name):
        """Fetch company reviews using search."""
        system_prompt = """You search and aggregate company reviews from Glassdoor, AmbitionBox, Reddit, etc.

Return ONLY valid JSON. No markdown, no extra text.

Format:
{"company_name":"","reviews":[{"source":"","rating":"","comment":"","url":""}],"aggregated_review_score":7,"summary":""}"""

        user_prompt = f"Company: {company_name}\n\nSearch reviews and return JSON only:"

        return self.send_request(system_prompt, user_prompt, use_search=True)


def get_html(url, timeout=15):
    """Fetch HTML from URL."""
    headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'}

    try:
        response = requests.get(url, headers=headers, timeout=timeout)
        response.raise_for_status()
        return str(BeautifulSoup(response.text, "html.parser"))
    except Exception as e:
        return f"ERROR: {e}"


def extract_json(text):
    """Extract JSON from response, handling markdown."""
    try:
        return json.loads(text)
    except:
        pass

    patterns = [r'```(?:json)?\s*\n(.*?)\n```', r'(\{.*\})']

    for pattern in patterns:
        matches = re.findall(pattern, text, re.DOTALL)
        for match in matches:
            try:
                return json.loads(match)
            except:
                continue

    print(f"    WARNING: JSON parse failed: {text[:150]}...")
    return {}


def read_urls_from_file(filepath):
    """Read URLs from txt or Excel file."""
    if filepath.endswith('.xlsx') or filepath.endswith('.xls'):
        try:
            df = pd.read_excel(filepath)
            url_col = next((col for col in df.columns if 'url' in col.lower()), df.columns[0])
            urls = df[url_col].dropna().tolist()
            print(f"[+] Loaded {len(urls)} URLs from Excel: {filepath}")
            return urls
        except Exception as e:
            print(f"[!] Excel read error: {e}")
            return []
    else:
        try:
            with open(filepath, 'r') as f:
                urls = [line.strip() for line in f if line.strip()]
            print(f"[+] Loaded {len(urls)} URLs from text file: {filepath}")
            return urls
        except Exception as e:
            print(f"[!] File read error: {e}")
            return []


def export_to_excel(job_data_list, review_data_list, output_filename=None):
    """Export job data and reviews to Excel."""
    if output_filename is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = f"jobs_{timestamp}.xlsx"

    wb = Workbook()
    wb.remove(wb.active)

    # Styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # Jobs Sheet
    if job_data_list:
        jobs_sheet = wb.create_sheet("Jobs Summary")

        jobs_rows = []
        for job in job_data_list:
            jobs_rows.append({
                "Company": job.get("company_name", ""),
                "Role": job.get("role_name", ""),
                "Experience": job.get("experience_required", ""),
                "Level": job.get("experience_type", ""),
                "Location": job.get("location", {}).get("city", ""),
                "Remote": job.get("remote", ""),
                "Match Score": job.get("match_score", 0)
            })

        jobs_rows = sorted(jobs_rows, key=lambda x: x.get("Match Score", 0), reverse=True)

        headers = ["Company", "Role", "Experience", "Level", "Location", "Remote", "Match Score"]
        for col_num, header in enumerate(headers, 1):
            cell = jobs_sheet.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = center_align

        for row_num, job_row in enumerate(jobs_rows, 2):
            for col_num, header in enumerate(headers, 1):
                cell = jobs_sheet.cell(row=row_num, column=col_num)
                cell.value = job_row.get(header, "")
                cell.border = border
                cell.alignment = left_align if header not in ["Match Score", "Remote"] else center_align

        jobs_sheet.column_dimensions['A'].width = 25
        jobs_sheet.column_dimensions['B'].width = 30
        jobs_sheet.column_dimensions['C'].width = 15
        jobs_sheet.column_dimensions['D'].width = 15
        jobs_sheet.column_dimensions['E'].width = 20
        jobs_sheet.column_dimensions['F'].width = 12
        jobs_sheet.column_dimensions['G'].width = 12

    # Reviews Sheet
    if review_data_list:
        reviews_sheet = wb.create_sheet("Company Reviews")

        headers = ["Company", "Review Score", "Source", "Rating", "Comment", "URL"]
        for col_num, header in enumerate(headers, 1):
            cell = reviews_sheet.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = center_align

        row_num = 2
        for review_company in review_data_list:
            company_name = review_company.get("company_name", "")
            aggregated_score = review_company.get("aggregated_review_score", 0)
            reviews = review_company.get("reviews", [])

            if not reviews:
                cell = reviews_sheet.cell(row=row_num, column=1)
                cell.value = company_name
                cell.border = border
                cell.alignment = left_align

                cell = reviews_sheet.cell(row=row_num, column=2)
                cell.value = aggregated_score
                cell.border = border
                cell.alignment = center_align

                row_num += 1
                continue

            for review in reviews:
                reviews_sheet.cell(row=row_num, column=1).value = company_name
                reviews_sheet.cell(row=row_num, column=1).border = border
                reviews_sheet.cell(row=row_num, column=1).alignment = left_align

                reviews_sheet.cell(row=row_num, column=2).value = aggregated_score
                reviews_sheet.cell(row=row_num, column=2).border = border
                reviews_sheet.cell(row=row_num, column=2).alignment = center_align

                reviews_sheet.cell(row=row_num, column=3).value = review.get("source", "")
                reviews_sheet.cell(row=row_num, column=3).border = border
                reviews_sheet.cell(row=row_num, column=3).alignment = left_align

                reviews_sheet.cell(row=row_num, column=4).value = review.get("rating", "")
                reviews_sheet.cell(row=row_num, column=4).border = border
                reviews_sheet.cell(row=row_num, column=4).alignment = center_align

                reviews_sheet.cell(row=row_num, column=5).value = review.get("comment", "")
                reviews_sheet.cell(row=row_num, column=5).border = border
                reviews_sheet.cell(row=row_num, column=5).alignment = left_align

                reviews_sheet.cell(row=row_num, column=6).value = review.get("url", "")
                reviews_sheet.cell(row=row_num, column=6).border = border
                reviews_sheet.cell(row=row_num, column=6).alignment = left_align

                row_num += 1

        reviews_sheet.column_dimensions['A'].width = 25
        reviews_sheet.column_dimensions['B'].width = 12
        reviews_sheet.column_dimensions['C'].width = 15
        reviews_sheet.column_dimensions['D'].width = 10
        reviews_sheet.column_dimensions['E'].width = 45
        reviews_sheet.column_dimensions['F'].width = 35

    # Summary Sheet
    summary_sheet = wb.create_sheet("Summary", 0)
    summary_sheet.column_dimensions['A'].width = 35
    summary_sheet.column_dimensions['B'].width = 20

    summary_data = [
        ["Job Scraping Report", ""],
        ["Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["", ""],
        ["Total Jobs Found", len(job_data_list)],
        ["Companies Reviewed", len(review_data_list)],
        ["", ""],
    ]

    if job_data_list:
        sorted_jobs = sorted(job_data_list, key=lambda x: x.get("match_score", 0), reverse=True)
        summary_data.append(["Top Matches", ""])
        for idx, job in enumerate(sorted_jobs[:5], 1):
            summary_data.append([
                f"{idx}. {job.get('company_name', '')} - {job.get('role_name', '')}",
                f"Score: {job.get('match_score', 0)}"
            ])

    for row_num, row_data in enumerate(summary_data, 1):
        for col_num, value in enumerate(row_data, 1):
            cell = summary_sheet.cell(row=row_num, column=col_num)
            cell.value = value

            if row_num in [1, 4, 7]:
                cell.font = Font(bold=True, size=12)

            if row_num == 1:
                cell.font = Font(bold=True, size=14, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    wb.save(output_filename)
    return output_filename


def main():
    """Main execution."""
    try:
        print("="*60)
        print("JOB FETCHER SCRIPT")
        print("="*60)

        # Initialize Gemini
        genai_client = Genai()
        print(f"[+] Gemini Model: {genai_client.model}")
        print(f"[+] API Key: {'Loaded' if genai_client.api_key else 'Missing'}")
        print()

        # Storage
        job_data_list = []
        review_data_list = []

        # Read URLs
        input_file = os.getenv('INPUT_FILE', 'job_links.txt')
        if os.path.exists('job_links.xlsx'):
            input_file = 'job_links.xlsx'

        urls = read_urls_from_file(input_file)

        if not urls:
            print("[!] No URLs found. Create 'job_links.txt' or 'job_links.xlsx'")
            return

        print(f"[+] Processing {len(urls)} job URLs\n")

        # Config
        config = {
            "preferred_locations": ["Bangalore", "Remote", "Hybrid"],
            "preferred_experience": "2-5 years",
            "preferred_skills": ["Python", "AI", "Backend"],
            "job_type": "Full-time"
        }

        # Process each URL
        for idx, url in enumerate(urls, 1):
            print(f"[{idx}/{len(urls)}] Processing: {url}")

            # Fetch HTML
            print("  -> Fetching HTML...")
            html_content = get_html(url)

            if html_content.startswith("ERROR"):
                print(f"  [!] {html_content}")
                continue

            print(f"  -> HTML fetched ({len(html_content)} chars)")

            # Parse job with Gemini
            print("  -> Extracting job data with Gemini...")
            job_response = genai_client.parse_job(html_content, config)

            if job_response.startswith("ERROR"):
                print(f"  [!] {job_response}")
                continue

            job_data = extract_json(job_response)

            if job_data:
                role = job_data.get('role_name', 'Unknown')
                company = job_data.get('company_name', 'Unknown')
                score = job_data.get('match_score', 0)
                print(f"  [+] Job: {role} at {company} (Match: {score}/10)")

                # Fetch reviews
                if company and company != 'Unknown':
                    print("  -> Fetching company reviews...")
                    review_response = genai_client.get_reviews(company)

                    if not review_response.startswith("ERROR"):
                        review_data = extract_json(review_response)

                        if review_data:
                            review_score = review_data.get('aggregated_review_score', 0)
                            print(f"  [+] Reviews: Score {review_score}/10")
                            review_data_list.append(review_data)

                job_data_list.append(job_data)
            else:
                print("  [!] Failed to extract job data")

            print()

        # Export to Excel
        print("="*60)
        if job_data_list or review_data_list:
            print("[+] Exporting to Excel...")
            output_file = export_to_excel(job_data_list, review_data_list)
            print(f"[+] SUCCESS! Excel file created: {output_file}")
        else:
            print("[!] No data to export")

        print(f"[+] Total jobs processed: {len(job_data_list)}")
        print(f"[+] Total companies reviewed: {len(review_data_list)}")
        print("="*60)

    except Exception as e:
        print(f"[!] Error: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()
